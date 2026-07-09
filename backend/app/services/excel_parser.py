"""Chunked Excel parsing with openpyxl read_only + data_only.

Safety:
- Never executes Remediation (or any cell) as a command.
- Streams rows to keep memory low for 20k+ record files.
"""

from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.constants.excel_columns import EXCEL_COLUMN_MAP, EXCEL_REQUIRED_COLUMNS


class ExcelColumnError(ValueError):
    """Raised when required Excel columns are missing or unmapped."""


@dataclass(frozen=True)
class ParsedExcelRow:
    row_number: int
    values: dict[str, str | None]


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def cell_to_str(value: Any) -> str | None:
    """Convert Excel cell values to plain strings without losing content."""
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return repr(value)
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    return text if text else None


def validate_headers(headers: list[str]) -> dict[int, str]:
    """Validate required columns and return {column_index: snake_case_field}."""
    normalized = [_normalize_header(h) for h in headers]
    present = {h for h in normalized if h}
    missing = [col for col in EXCEL_REQUIRED_COLUMNS if col not in present]
    if missing:
        raise ExcelColumnError(
            "Missing required Excel columns: " + ", ".join(missing)
        )

    index_to_field: dict[int, str] = {}
    for idx, header in enumerate(normalized):
        if not header:
            continue
        field = EXCEL_COLUMN_MAP.get(header)
        if field:
            index_to_field[idx] = field
    return index_to_field


def _row_is_empty(values: dict[str, str | None]) -> bool:
    return all(v is None or str(v).strip() == "" for v in values.values())


def iter_excel_rows(
    file_path: str | Path,
    *,
    chunk_size: int = 1000,
) -> Iterator[list[ParsedExcelRow]]:
    """Yield chunks of normalized snake_case rows (openpyxl read_only streaming)."""
    path = Path(file_path)
    if not path.is_file():
        raise FileNotFoundError(f"Excel file not found: {path}")

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)

        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise ExcelColumnError("Excel file is empty") from exc

        headers = [_normalize_header(cell) for cell in header_row]
        index_to_field = validate_headers(headers)

        chunk: list[ParsedExcelRow] = []
        excel_row_number = 1  # header is row 1

        for raw_row in rows_iter:
            excel_row_number += 1
            values: dict[str, str | None] = {
                field: None for field in EXCEL_COLUMN_MAP.values()
            }
            for idx, field in index_to_field.items():
                cell_value = raw_row[idx] if idx < len(raw_row) else None
                values[field] = cell_to_str(cell_value)

            if _row_is_empty(values):
                continue

            chunk.append(ParsedExcelRow(row_number=excel_row_number, values=values))
            if len(chunk) >= chunk_size:
                yield chunk
                chunk = []

        if chunk:
            yield chunk
    finally:
        workbook.close()
