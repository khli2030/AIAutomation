"""Phase 6.5 — full end-to-end mock workflow (MOCK_MODE=true).

Covers: seed → Excel upload → parse → validate → READY_FOR_PLAN → plan →
waiting_dry_run → mock dry-run → dry_run results → approve → mock run →
run results → final job status. Dashboard 501 is accepted when unimplemented.

Never calls ansible-runner, ansible-playbook, subprocess Ansible, or SSH.
"""

from __future__ import annotations

import ast
import sys
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.constants.excel_columns import EXCEL_REQUIRED_COLUMNS
from app.constants.import_status import ImportBatchStatus
from app.constants.job_result_type import JobResultType
from app.constants.job_status import JobStatus
from app.constants.record_status import RecordStatus
from app.db.seed_assets import MVP_TEST_ASSETS

FINAL_JOB_STATUSES = {
    JobStatus.SUCCESS.value,
    JobStatus.FAILED.value,
    JobStatus.PARTIALLY_FAILED.value,
}

FORBIDDEN_IMPORT_ROOTS = {
    "ansible_runner",
    "subprocess",
    "paramiko",
    "fabric",
    "invoke",
}


def _col(name: str) -> int:
    return list(EXCEL_REQUIRED_COLUMNS).index(name)


def build_sample_compliance_xlsx() -> bytes:
    """Build a small Qualys/MBSS-style workbook with READY_FOR_PLAN-eligible rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance"
    ws.append(list(EXCEL_REQUIRED_COLUMNS))

    # Two hosts, same enabled task (PermitRootLogin) — indices 0 and 1 → both success in mock.
    devices = [
        ("e2e-linux-01", "10.20.0.11", "CTRL-ROOT-01"),
        ("e2e-linux-02", "10.20.0.12", "CTRL-ROOT-02"),
    ]
    for device_name, _ip, control_id in devices:
        row = [""] * len(EXCEL_REQUIRED_COLUMNS)
        row[_col("Device Name")] = device_name
        row[_col("Overall Status")] = "Failed"
        row[_col("Criticality")] = "High"
        row[_col("Qualys Control ID")] = control_id
        row[_col("Source Check ID")] = f"SRC-{control_id}"
        row[_col("Control Description")] = "SSH PermitRootLogin must be no"
        row[_col("RATIONALE")] = "Prevent direct root SSH access"
        row[_col("Remediation")] = (
            "Set PermitRootLogin no in sshd_config (stored text only — never executed)"
        )
        row[_col("Expected Configuration")] = "PermitRootLogin no"
        row[_col("Config Scan ID")] = f"SCAN-{control_id}"
        ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _imported_roots(module_file: Path) -> set[str]:
    tree = ast.parse(module_file.read_text(encoding="utf-8"))
    imported: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                imported.add(alias.name.split(".")[0])
        elif isinstance(node, ast.ImportFrom) and node.module:
            parts = node.module.split(".")
            imported.add(parts[0])
            imported.add(parts[-1])
    return imported


@pytest.fixture(autouse=True)
def _clear_real_adapter_import() -> None:
    sys.modules.pop("app.services.real_ansible_runner", None)
    yield
    sys.modules.pop("app.services.real_ansible_runner", None)


def test_e2e_mock_workflow_full_path(
    e2e_client: TestClient,
    auth_headers: dict[str, str],
    e2e_env: dict,
) -> None:
    """Drive the complete mock remediation path via HTTP API."""
    assert e2e_env["settings"].mock_mode is True

    # --- 1–2. Seeds already applied in fixture; confirm via DB ---
    session_factory = e2e_env["session_factory"]
    db = session_factory()
    try:
        from app.models.asset import Asset
        from app.models.remediation_catalog import RemediationCatalog

        assets = db.query(Asset).filter(Asset.is_active.is_(True)).all()
        assert len(assets) >= len(MVP_TEST_ASSETS)
        enabled = (
            db.query(RemediationCatalog)
            .filter(RemediationCatalog.is_enabled.is_(True))
            .all()
        )
        assert any(c.task_code == "SSH_DISABLE_ROOT_LOGIN" for c in enabled)
        assert all(c.task_code == "SSH_DISABLE_ROOT_LOGIN" for c in enabled)
    finally:
        db.close()

    # --- 3. Upload sample Excel ---
    xlsx_bytes = build_sample_compliance_xlsx()
    upload = e2e_client.post(
        "/imports/upload",
        headers=auth_headers,
        files={
            "file": (
                "e2e_compliance.xlsx",
                xlsx_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"uploaded_by": "e2e-operator"},
    )
    assert upload.status_code == 202, upload.text
    batch_id = upload.json()["batch"]["id"]
    assert batch_id > 0

    # --- 4. Parse (synchronous via fixture patch of Celery .delay) ---
    batch = e2e_client.get(f"/imports/{batch_id}", headers=auth_headers)
    assert batch.status_code == 200, batch.text
    batch_body = batch.json()
    assert batch_body["status"] == ImportBatchStatus.PARSED.value
    assert batch_body["total_records"] == 2
    assert batch_body["valid_records"] == 2
    assert batch_body["invalid_records"] == 0

    # --- 5. Validate ---
    validate = e2e_client.post(f"/imports/{batch_id}/validate", headers=auth_headers)
    assert validate.status_code == 200, validate.text
    summary = validate.json()
    assert summary["total_records"] == 2
    assert summary["ready_for_plan"] == 2
    assert summary["asset_not_found"] == 0
    assert summary["needs_review"] == 0

    # --- 6. Confirm READY_FOR_PLAN records ---
    records = e2e_client.get(f"/imports/{batch_id}/records", headers=auth_headers)
    assert records.status_code == 200, records.text
    items = records.json()["items"]
    assert len(items) == 2
    for item in items:
        assert item["validation_status"] == RecordStatus.READY_FOR_PLAN.value
        assert item["task_code"] == "SSH_DISABLE_ROOT_LOGIN"

    # --- 7. Generate execution plan ---
    plan_resp = e2e_client.post(
        f"/imports/{batch_id}/generate-plan",
        headers=auth_headers,
        params={"created_by": "e2e-operator"},
    )
    assert plan_resp.status_code == 200, plan_resp.text
    plan = plan_resp.json()["plan"]
    plan_id = plan["id"]
    assert plan["job_count"] >= 1
    assert plan["target_count"] == 2
    assert plan["ready_for_plan_records"] == 2
    assert plan["skipped_disabled_catalog"] == 0

    # --- 8. Confirm jobs waiting_dry_run ---
    jobs_resp = e2e_client.get(f"/execution-plans/{plan_id}/jobs", headers=auth_headers)
    assert jobs_resp.status_code == 200, jobs_resp.text
    jobs = jobs_resp.json()["items"]
    assert len(jobs) >= 1
    for job in jobs:
        assert job["status"] == JobStatus.WAITING_DRY_RUN.value
        assert job["task_code"] == "SSH_DISABLE_ROOT_LOGIN"
        assert job["target_count"] >= 1

    # Process each job through dry-run → approve → run
    for job in jobs:
        job_id = job["id"]

        # --- 9. Mock dry-run ---
        dry = e2e_client.post(f"/execution-jobs/{job_id}/dry-run", headers=auth_headers)
        assert dry.status_code == 200, dry.text
        dry_body = dry.json()
        assert dry_body["mock_mode"] is True
        assert dry_body["mode"] == "dry_run"
        assert dry_body["status"] == JobStatus.DRY_RUN_SUCCESS.value
        assert dry_body["hosts_total"] == job["target_count"]
        assert dry_body["hosts_failed"] == 0

        # --- 10. Confirm dry_run results ---
        dry_results = e2e_client.get(
            f"/execution-jobs/{job_id}/results",
            headers=auth_headers,
            params={"result_type": "dry_run"},
        )
        assert dry_results.status_code == 200, dry_results.text
        dry_payload = dry_results.json()
        assert dry_payload["total"] == job["target_count"]
        assert dry_payload["result_type_filter"] == JobResultType.DRY_RUN.value
        assert all(
            r["result_type"] == JobResultType.DRY_RUN.value for r in dry_payload["items"]
        )
        assert all(r["status"] in {"success", "skipped"} for r in dry_payload["items"])

        # --- 11. Approve after dry_run_success ---
        approve = e2e_client.post(
            f"/execution-jobs/{job_id}/approve",
            headers=auth_headers,
            json={"reviewed_by": "e2e-approver"},
        )
        assert approve.status_code == 200, approve.text
        assert approve.json()["status"] == JobStatus.APPROVED.value
        assert approve.json()["approved_by"] == "e2e-approver"

        # --- 12. Mock execution ---
        run = e2e_client.post(f"/execution-jobs/{job_id}/run", headers=auth_headers)
        assert run.status_code == 200, run.text
        run_body = run.json()
        assert run_body["mock_mode"] is True
        assert run_body["mode"] == "apply"
        assert run_body["status"] in FINAL_JOB_STATUSES
        assert run_body["hosts_total"] == job["target_count"]

        # --- 13. Confirm run results ---
        run_results = e2e_client.get(
            f"/execution-jobs/{job_id}/results",
            headers=auth_headers,
            params={"result_type": "run"},
        )
        assert run_results.status_code == 200, run_results.text
        run_payload = run_results.json()
        assert run_payload["total"] == job["target_count"]
        assert run_payload["result_type_filter"] == JobResultType.RUN.value
        assert all(
            r["result_type"] == JobResultType.RUN.value for r in run_payload["items"]
        )

        # Dry-run rows must still exist (run must not overwrite them).
        both = e2e_client.get(
            f"/execution-jobs/{job_id}/results",
            headers=auth_headers,
        )
        assert both.status_code == 200, both.text
        both_items = both.json()["items"]
        dry_count = sum(
            1 for r in both_items if r["result_type"] == JobResultType.DRY_RUN.value
        )
        run_count = sum(
            1 for r in both_items if r["result_type"] == JobResultType.RUN.value
        )
        assert dry_count == job["target_count"]
        assert run_count == job["target_count"]

        # --- 14. Final job status ---
        assert run_payload["job_status"] in FINAL_JOB_STATUSES

    # --- 15. Dashboard summary (501 until Phase 7 is fine) ---
    dash = e2e_client.get("/dashboard/summary", headers=auth_headers)
    assert dash.status_code in {200, 501}, dash.text
    if dash.status_code == 200:
        body = dash.json()
        assert isinstance(body, dict)


def test_e2e_mock_workflow_rejects_unauthenticated(e2e_client: TestClient) -> None:
    resp = e2e_client.get("/imports/1")
    assert resp.status_code in {401, 503}


def test_e2e_execution_modules_stay_mock_safe() -> None:
    """Static guard: ansible_execution must not import runner/subprocess/SSH libs."""
    root = Path(__file__).resolve().parents[2] / "app" / "services"
    for name in ("ansible_execution.py", "job_approval.py", "plan_generator.py"):
        imported = _imported_roots(root / name)
        assert not (imported & FORBIDDEN_IMPORT_ROOTS), (
            f"{name} imports forbidden modules: {imported & FORBIDDEN_IMPORT_ROOTS}"
        )


def test_build_sample_xlsx_has_required_headers() -> None:
    from openpyxl import load_workbook

    data = build_sample_compliance_xlsx()
    wb = load_workbook(BytesIO(data), read_only=True)
    headers = [c.value for c in next(wb.active.iter_rows(min_row=1, max_row=1))]
    assert headers == list(EXCEL_REQUIRED_COLUMNS)
    rows = list(wb.active.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
